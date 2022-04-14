import openpyxl


class HomePageData:
    test_homepage_data = [{"full_name": 'standard_user', "password": 'secret_sauce'}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\91983\\Documents\\Book1.xlsx")  # Excel file path
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get rows
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(f"this is the dictionary {Dict}")
        return [Dict]
